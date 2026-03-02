from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


@dataclass(frozen=True)
class BookingCell:
  raw_value: str
  status: str


def _clean_text(value: object) -> str:
  if value is None:
    return ""
  return str(value).strip()


def _normalize_slot(slot_value: object) -> tuple[str, int]:
  cleaned = _clean_text(slot_value).replace("'", "")
  match = re.match(r"^(.*?)[\.-](\d+)$", cleaned)
  if not match:
    raise ValueError(f"Cannot parse slot label: {slot_value!r}")
  prefix = match.group(1).strip().upper()
  slot_no = int(match.group(2))
  return prefix, slot_no


def _classify_booking_cell(value: object) -> BookingCell:
  raw = _clean_text(value)
  upper = raw.upper()
  if not raw or raw == "0":
    return BookingCell(raw_value=raw, status="empty")
  if "THU HỒI" in upper or "KHAI THÁC" in upper or "OFF" in upper or "NGHỈ" in upper:
    return BookingCell(raw_value=raw, status="blocked")
  return BookingCell(raw_value=raw, status="booked")


def _safe_key(value: object) -> str:
  return _clean_text(value).upper().replace(" ", "")


def _iter_data_rows(sheet) -> Iterable[int]:
  for row_idx in range(6, sheet.max_row + 1):
    bd_code = _clean_text(sheet.cell(row_idx, 2).value)
    slot_label = _clean_text(sheet.cell(row_idx, 22).value)
    code_report = _clean_text(sheet.cell(row_idx, 16).value)
    if bd_code and slot_label and code_report:
      yield row_idx


def build_outputs(source_xlsx: Path, out_dir: Path) -> dict[str, object]:
  wb = load_workbook(source_xlsx, data_only=False)
  ws = wb[wb.sheetnames[0]]

  week_meta = {
    "W08": _clean_text(ws.cell(3, 26).value),
    "W09": _clean_text(ws.cell(3, 27).value),
  }

  screen_records: dict[tuple[str, str], dict[str, object]] = {}
  slot_records: list[dict[str, object]] = []

  for row_idx in _iter_data_rows(ws):
    bd_code = _clean_text(ws.cell(row_idx, 2).value)
    city = _clean_text(ws.cell(row_idx, 4).value)
    district = _clean_text(ws.cell(row_idx, 5).value)
    address = _clean_text(ws.cell(row_idx, 6).value)
    traffic = ws.cell(row_idx, 10).value
    channel = _clean_text(ws.cell(row_idx, 15).value)
    code_report = _clean_text(ws.cell(row_idx, 16).value)
    location_name = _clean_text(ws.cell(row_idx, 18).value)
    screen_type = _clean_text(ws.cell(row_idx, 19).value)
    site = _clean_text(ws.cell(row_idx, 20).value)
    model = _clean_text(ws.cell(row_idx, 21).value)

    slot_prefix, slot_no = _normalize_slot(ws.cell(row_idx, 22).value)
    slot_label = f"{slot_prefix}.{slot_no}"
    screen_key = (code_report.upper(), slot_prefix)
    screen_uid = f"{_safe_key(code_report)}__{slot_prefix}"

    if screen_key not in screen_records:
      screen_records[screen_key] = {
        "screen_uid": screen_uid,
        "code_report": code_report,
        "slot_prefix": slot_prefix,
        "bd_code": bd_code,
        "channel": channel,
        "city": city,
        "district": district,
        "address": address,
        "location_name": location_name,
        "screen_type": screen_type,
        "site": site,
        "model": model,
        "weekly_traffic": int(traffic) if isinstance(traffic, (int, float)) else 0,
      }

    for week_code, col in [("W08", 26), ("W09", 27)]:
      cell = _classify_booking_cell(ws.cell(row_idx, col).value)
      slot_records.append(
        {
          "screen_uid": screen_uid,
          "code_report": code_report,
          "slot_prefix": slot_prefix,
          "slot_no": slot_no,
          "slot_label": slot_label,
          "week_code": week_code,
          "week_range": week_meta.get(week_code, ""),
          "booking_value": cell.raw_value,
          "booking_status": cell.status,
          "is_booked": 1 if cell.status == "booked" else 0,
        }
      )

  inventory_df = pd.DataFrame(screen_records.values()).sort_values(["code_report", "slot_prefix"])
  slots_df = pd.DataFrame(slot_records).sort_values(["code_report", "slot_prefix", "slot_no", "week_code"])

  slot_check = (
    slots_df.groupby(["screen_uid", "week_code"])["slot_no"]
    .nunique()
    .reset_index(name="slot_count")
  )
  bad_screens = slot_check[slot_check["slot_count"] != 32]

  out_dir.mkdir(parents=True, exist_ok=True)
  inventory_path = out_dir / "booking_w09_screen_inventory.csv"
  slots_path = out_dir / "booking_w09_slot_week_status.csv"
  summary_path = out_dir / "booking_w09_summary.json"

  inventory_df.to_csv(inventory_path, index=False, encoding="utf-8-sig")
  slots_df.to_csv(slots_path, index=False, encoding="utf-8-sig")

  summary = {
    "source_file": str(source_xlsx),
    "sheet_name": ws.title,
    "total_data_rows": int(len(list(_iter_data_rows(ws)))),
    "unique_screens": int(inventory_df.shape[0]),
    "total_slot_week_rows": int(slots_df.shape[0]),
    "weeks": week_meta,
    "booked_rows": int((slots_df["booking_status"] == "booked").sum()),
    "blocked_rows": int((slots_df["booking_status"] == "blocked").sum()),
    "empty_rows": int((slots_df["booking_status"] == "empty").sum()),
    "slot_integrity_ok": bool(bad_screens.empty),
    "slot_integrity_issues": bad_screens.to_dict(orient="records"),
    "output_files": {
      "screen_inventory_csv": str(inventory_path),
      "slot_week_status_csv": str(slots_path),
    },
  }
  summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
  return summary


def main() -> None:
  parser = argparse.ArgumentParser(description="Import booking progress Excel into normalized CSVs.")
  parser.add_argument("--input", required=True, help="Path to booking progress .xlsx file")
  parser.add_argument(
    "--output-dir",
    default="output/spreadsheet",
    help="Directory for generated CSV and summary files",
  )
  args = parser.parse_args()

  source_xlsx = Path(args.input)
  out_dir = Path(args.output_dir)
  summary = build_outputs(source_xlsx, out_dir)
  print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
  main()
